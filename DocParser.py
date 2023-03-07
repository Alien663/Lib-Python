import os
import csv
import openpyxl

class MyExcel():
    def __init__(self, filename, header=1):
        self.filename = filename
        self.header = header
        self.__extend = self.filename.split(".")[-1]
        if(self.__extend not in ["xlsx", "xls", "csv"]):
            raise Exception("Can not know this file with extend file name : ." + self.__extend)
    
    def __str__(self):
        return "read data from %s" % self.filename

    def read(self):
        match self.__extend:
            case "xlsx":
                return self.__read_excel()
            case "xls":
                return self.__read_excel()
            case "csv":
                return self.__read_csv()

    def __read_excel(self):
        result = dict()
        workbook = openpyxl.load_workbook(self.filename)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            temp_sheet = []
            headers = ["nothing"]
            for i in range(1, sheet.max_column + 1):
                headers.append(sheet.cell(self.header, i).value)
            
            for i in range(1, sheet.max_row + 1):
                temp_row = dict()
                for j in range(1, sheet.max_column + 1):
                    temp_row[headers[j]] = sheet.cell(i, j).value
                temp_sheet.append(temp_row)
        result[sheetname] = temp_sheet
        return result

    def __read_csv(self):
        result = []
        with open(self.filename, 'r', encoding='utf-8', newline='') as fp:
            rows = csv.DictReader(fp)
            for row in rows:
                result.append(row)
        return(result)

    def export(self, data):
        match self.__extend:
            case "xlsx":
                self.__export_excel(data)
            case "xls":
                self.__export_excel(data)
            case "csv":
                self.__export_csv(data)
    
    def __export_excel(self, data):
        workbook = openpyxl.Workbook()
        the_first = True
        for sheet in data:
            if(the_first):
                temp_sheet = workbook["Sheet"]
                temp_sheet.title = sheet
                the_first = False
            else:
                temp_sheet = workbook.create_sheet(sheet)
            j=1
            for colname in data[sheet][self.header]:
                temp_sheet.cell(self.header, j).value = colname
                j+=1

            for i in range(self.header + 1, len(data[sheet]) + 1):
                temp_row = data[sheet][i-1]
                j=1
                for colname in temp_row:
                    temp_sheet.cell(i, j).value = temp_row[colname]
                    j+=1
        workbook.save(self.filename)

    def __export_csv(self, data):
        with open(self.filename, 'w', encoding="utf-8") as fd:
            writer = csv.DictWriter(fd, list(data[0].keys()))
            writer.writeheader()
            for row in data:
                writer.writerow(row)
        fd.close()